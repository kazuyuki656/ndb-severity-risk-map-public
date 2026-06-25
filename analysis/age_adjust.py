"""年齢・性別調整版スコア（直接法標準化）

各二次医療圏のリスク該当率を、全国の特定健診受診者の年齢×性別構成を
基準集団として直接法で標準化する。これにより、地域間の年齢構成の違い
（高齢化地域でリスクが高く出る交絡）を除去する。

直接法: 調整率 = Σ_層 (層別リスク率_地域 × 標準集団の層の重み)
  層 = 7年齢階級 × 2性別 = 14層
  標準集団 = 全圏域合計の層別受診者数
"""
import os
import openpyxl
import pandas as pd
import numpy as np
from config import DATA_DIR, OUTPUT_DIR, REGION_FILES, RISK_THRESHOLDS

# 列インデックス（0始まり）。男: 年齢40-44..70-74 = col4..10, 中計11。女: col12..18, 中計19。
MALE_AGE_COLS = list(range(4, 11))     # 7列
FEMALE_AGE_COLS = list(range(12, 19))  # 7列
AGE_LABELS = ['40-44', '45-49', '50-54', '55-59', '60-64', '65-69', '70-74']


def safe_int(v):
    if v is None or v in ('-', '‐', '－', ''):
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def parse_region_strata(filepath: str) -> list[dict]:
    """二次医療圏xlsxを 層別（性×年齢）人数つきの行リストに変換。"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    cur_pref = cur_code = cur_name = None

    for row in rows[5:]:
        if row[0] is not None:
            cur_pref = str(row[0]).strip()
        if row[1] is not None:
            cur_code = str(row[1]).strip()
        if row[2] is not None:
            cur_name = str(row[2]).strip()
        category = row[3]
        if category is None:
            continue
        category = str(category).strip()
        if cur_pref in ('二次医療圏判別不可',):
            continue

        rec = {'pref': cur_pref, 'region_code': cur_code,
               'region_name': cur_name, 'category': category}
        for sex, cols in (('M', MALE_AGE_COLS), ('F', FEMALE_AGE_COLS)):
            for age_label, c in zip(AGE_LABELS, cols):
                rec[f'{sex}_{age_label}'] = safe_int(row[c]) if c < len(row) else 0
        records.append(rec)

    return records


def build_strata_dataframe(indicator: str, records: list[dict]) -> pd.DataFrame:
    """層別人数を long 形式（pref×region×category×stratum→count）に展開。"""
    df = pd.DataFrame(records)
    stratum_cols = [f'{s}_{a}' for s in ('M', 'F') for a in AGE_LABELS]
    long = df.melt(
        id_vars=['pref', 'region_code', 'region_name', 'category'],
        value_vars=stratum_cols, var_name='stratum', value_name='n',
    )
    long['indicator'] = indicator
    return long


def compute_adjusted_rates(indicator: str, long: pd.DataFrame) -> pd.DataFrame:
    """直接法による年齢・性別調整リスク率を圏域別に算出。"""
    threshold = RISK_THRESHOLDS[indicator]
    risk_cats = set(threshold['risk_categories'])

    long = long.copy()
    long['is_risk'] = long['category'].isin(risk_cats)

    # --- 層別の分母（受診者数）: 各 pref×region×stratum のカテゴリ合計 ---
    denom = long.groupby(['pref', 'region_code', 'region_name', 'stratum'])['n'].sum().reset_index()
    denom = denom.rename(columns={'n': 'denom'})

    # --- 層別の分子（リスク該当数） ---
    numer = long[long['is_risk']].groupby(
        ['pref', 'region_code', 'region_name', 'stratum'])['n'].sum().reset_index()
    numer = numer.rename(columns={'n': 'numer'})

    strata = denom.merge(numer, on=['pref', 'region_code', 'region_name', 'stratum'], how='left')
    strata['numer'] = strata['numer'].fillna(0)
    strata['rate'] = np.where(strata['denom'] > 0, strata['numer'] / strata['denom'], np.nan)

    # --- 標準集団: 全圏域合計の層別受診者数 → 重み ---
    std_pop = strata.groupby('stratum')['denom'].sum()
    std_weight = (std_pop / std_pop.sum()).to_dict()
    strata['std_w'] = strata['stratum'].map(std_weight)

    # --- 直接法: 調整率 = Σ(層率 × 標準重み)。層率欠損(分母0)は標準重み再正規化で除外 ---
    results = []
    for (pref, code, name), g in strata.groupby(['pref', 'region_code', 'region_name']):
        valid = g.dropna(subset=['rate'])
        if valid['std_w'].sum() > 0:
            w = valid['std_w'] / valid['std_w'].sum()
            adj_rate = float((valid['rate'] * w).sum())
        else:
            adj_rate = np.nan
        crude_num = g['numer'].sum()
        crude_den = g['denom'].sum()
        crude_rate = crude_num / crude_den if crude_den > 0 else np.nan
        results.append({
            'pref': pref, 'region_code': code, 'region_name': name,
            'indicator': indicator, 'domain': threshold['domain'],
            'n_total': int(crude_den), 'crude_rate': crude_rate,
            'adj_rate': adj_rate,
        })

    return pd.DataFrame(results)


def run_age_adjustment():
    all_adj = []
    for indicator, filename in REGION_FILES.items():
        filepath = os.path.join(DATA_DIR, filename)
        print(f"Age-adjusting: {indicator} ({filename})")
        records = parse_region_strata(filepath)
        long = build_strata_dataframe(indicator, records)
        adj = compute_adjusted_rates(indicator, long)
        d = (adj['adj_rate'] - adj['crude_rate']).abs().mean()
        print(f"  {len(adj)} regions | crude {adj['crude_rate'].mean():.4f} "
              f"-> adj {adj['adj_rate'].mean():.4f} | mean|Δ| {d:.4f}")
        all_adj.append(adj)

    combined = pd.concat(all_adj, ignore_index=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    tidy_path = os.path.join(OUTPUT_DIR, 'region_adjusted_rates.csv')
    combined.to_csv(tidy_path, index=False, encoding='utf-8-sig')
    print(f"\nSaved: {tidy_path} ({len(combined)} rows)")
    return combined


def score_adjusted(combined: pd.DataFrame) -> pd.DataFrame:
    """調整率を偏差値化 → 領域スコア → 複合スコア（score.pyと同一手法）。"""
    df = combined.copy()
    df['z'] = df.groupby('indicator')['adj_rate'].transform(lambda x: (x - x.mean()) / x.std())
    df['dev'] = df['z'] * 10 + 50

    domain = df.groupby(['pref', 'region_code', 'region_name', 'domain'])['dev'].mean().reset_index()
    domain = domain.rename(columns={'dev': 'domain_score'})

    comp = domain.pivot_table(index=['pref', 'region_code', 'region_name'],
                              columns='domain', values='domain_score').reset_index()
    comp.columns.name = None
    domains = [d for d in ['糖尿病', 'CKD', '心血管'] if d in comp.columns]
    comp['composite_score'] = comp[domains].mean(axis=1)
    comp['risk_rank'] = comp['composite_score'].rank(ascending=False, method='min').astype(int)
    comp['risk_tier'] = pd.qcut(comp['composite_score'], 5,
                                labels=['低リスク', 'やや低', '中程度', 'やや高', '高リスク'])
    comp = comp.sort_values('composite_score', ascending=False)

    path = os.path.join(OUTPUT_DIR, 'composite_scores_adjusted.csv')
    comp.to_csv(path, index=False, encoding='utf-8-sig')
    print(f"Saved: {path}")
    return comp


def compare_crude_vs_adjusted(comp_adj: pd.DataFrame):
    """粗スコアと調整スコアの順位変動を比較。"""
    crude = pd.read_csv(os.path.join(OUTPUT_DIR, 'composite_scores.csv'), encoding='utf-8-sig')
    crude = crude[['region_code', 'pref', 'region_name', 'composite_score', 'risk_rank']]
    crude = crude.rename(columns={'composite_score': 'crude_score', 'risk_rank': 'crude_rank'})
    adj = comp_adj[['region_code', 'composite_score', 'risk_rank']].rename(
        columns={'composite_score': 'adj_score', 'risk_rank': 'adj_rank'})

    crude['region_code'] = crude['region_code'].astype(str).str.zfill(4)
    adj['region_code'] = adj['region_code'].astype(str).str.zfill(4)

    m = crude.merge(adj, on='region_code', how='inner')
    m['rank_change'] = m['crude_rank'] - m['adj_rank']  # 正=調整で順位上昇(悪化)
    m = m.sort_values('adj_rank')

    path = os.path.join(OUTPUT_DIR, 'crude_vs_adjusted.csv')
    m.to_csv(path, index=False, encoding='utf-8-sig')

    r = m[['crude_rank', 'adj_rank']].corr(method='spearman').iloc[0, 1]
    out = os.path.join(OUTPUT_DIR, 'adjustment_summary.txt')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(f"粗スコアと年齢調整スコアの順位相関(Spearman): {r:.3f}\n\n")
        f.write(f"順位変動の大きい圏域（年齢調整で順位が大きく下落＝粗スコアは年齢構成で過大評価）\n")
        for _, x in m.nlargest(10, 'rank_change').iterrows():
            f.write(f"  {x['pref']}/{x['region_name']}: "
                    f"{int(x['crude_rank'])}位 → {int(x['adj_rank'])}位 "
                    f"(Δ{int(x['rank_change']):+d})\n")
        f.write(f"\n順位変動の大きい圏域（年齢調整で順位が大きく上昇＝粗スコアは年齢構成で過小評価）\n")
        for _, x in m.nsmallest(10, 'rank_change').iterrows():
            f.write(f"  {x['pref']}/{x['region_name']}: "
                    f"{int(x['crude_rank'])}位 → {int(x['adj_rank'])}位 "
                    f"(Δ{int(x['rank_change']):+d})\n")
        f.write(f"\n=== 年齢調整後 複合スコア上位15圏域 ===\n")
        for _, x in m.head(15).iterrows():
            f.write(f"  {int(x['adj_rank']):3d}. {x['pref']}/{x['region_name']}  "
                    f"調整={x['adj_score']:.1f} (粗={x['crude_score']:.1f})\n")
    print(f"Saved: {path}, {out}")
    print(f"順位相関(Spearman) crude vs adjusted: {r:.3f}")
    return m


if __name__ == '__main__':
    combined = run_age_adjustment()
    comp_adj = score_adjusted(combined)
    compare_crude_vs_adjusted(comp_adj)
