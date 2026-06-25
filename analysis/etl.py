"""NDBオープンデータ ETL: xlsxから二次医療圏別tidy形式CSVへ変換"""
import os
import openpyxl
import pandas as pd
from config import (DATA_DIR, OUTPUT_DIR, REGION_FILES, RISK_THRESHOLDS,
                    MALE_SUBTOTAL, FEMALE_SUBTOTAL)


def parse_region_file(filepath: str) -> list[dict]:
    """二次医療圏別xlsxを行リストに変換。マージセルを前方充填で処理。"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    data_rows = rows[5:]  # ヘッダ5行をスキップ

    records = []
    current_pref = None
    current_region_code = None
    current_region_name = None

    for row in data_rows:
        pref = row[0]
        region_code = row[1]
        region_name = row[2]
        category = row[3]

        if pref is not None:
            current_pref = str(pref).strip()
        if region_code is not None:
            current_region_code = str(region_code).strip()
        if region_name is not None:
            current_region_name = str(region_name).strip()

        if category is None:
            continue

        category = str(category).strip()

        if current_pref in ('二次医療圏判別不可',):
            continue

        def safe_int(v):
            if v is None or v == '-' or v == '‐' or v == '－':
                return 0
            try:
                return int(v)
            except (ValueError, TypeError):
                return 0

        male_subtotal = safe_int(row[MALE_SUBTOTAL])
        female_subtotal = safe_int(row[FEMALE_SUBTOTAL])

        records.append({
            'pref': current_pref,
            'region_code': current_region_code,
            'region_name': current_region_name,
            'category': category,
            'male': male_subtotal,
            'female': female_subtotal,
            'total': male_subtotal + female_subtotal,
        })

    return records


def compute_risk_rates(indicator: str, records: list[dict]) -> pd.DataFrame:
    """指標ごとにリスク該当率を二次医療圏別に計算。"""
    threshold = RISK_THRESHOLDS[indicator]
    risk_cats = set(threshold['risk_categories'])

    df = pd.DataFrame(records)

    grouped = df.groupby(['pref', 'region_code', 'region_name'])

    results = []
    for (pref, code, name), group in grouped:
        total_all = group['total'].sum()
        risk_count = group[group['category'].isin(risk_cats)]['total'].sum()

        male_all = group['male'].sum()
        male_risk = group[group['category'].isin(risk_cats)]['male'].sum()

        female_all = group['female'].sum()
        female_risk = group[group['category'].isin(risk_cats)]['female'].sum()

        results.append({
            'pref': pref,
            'region_code': code,
            'region_name': name,
            'indicator': indicator,
            'domain': threshold['domain'],
            'n_total': total_all,
            'n_risk': risk_count,
            'risk_rate': risk_count / total_all if total_all > 0 else None,
            'n_male': male_all,
            'n_risk_male': male_risk,
            'risk_rate_male': male_risk / male_all if male_all > 0 else None,
            'n_female': female_all,
            'n_risk_female': female_risk,
            'risk_rate_female': female_risk / female_all if female_all > 0 else None,
        })

    return pd.DataFrame(results)


def run_etl():
    """全指標をETL処理し、統合CSVを出力。"""
    all_results = []

    for indicator, filename in REGION_FILES.items():
        filepath = os.path.join(DATA_DIR, filename)
        print(f"Processing: {indicator} ({filename})")

        records = parse_region_file(filepath)
        print(f"  Parsed {len(records)} rows")

        risk_df = compute_risk_rates(indicator, records)
        print(f"  {len(risk_df)} regions, "
              f"mean risk rate: {risk_df['risk_rate'].mean():.4f}")

        all_results.append(risk_df)

    combined = pd.concat(all_results, ignore_index=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    tidy_path = os.path.join(OUTPUT_DIR, 'region_risk_rates.csv')
    combined.to_csv(tidy_path, index=False, encoding='utf-8-sig')
    print(f"\nSaved tidy data: {tidy_path} ({len(combined)} rows)")

    pivot = combined.pivot_table(
        index=['pref', 'region_code', 'region_name'],
        columns='indicator',
        values='risk_rate',
    ).reset_index()

    pivot_path = os.path.join(OUTPUT_DIR, 'region_risk_pivot.csv')
    pivot.to_csv(pivot_path, index=False, encoding='utf-8-sig')
    print(f"Saved pivot data: {pivot_path} ({len(pivot)} rows)")

    return combined


if __name__ == '__main__':
    run_etl()
